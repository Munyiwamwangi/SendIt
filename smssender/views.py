from django.shortcuts import render
import openpyxl
import africastalking


def index(request):
    if "GET" == request.method:
        return render(request, 'smssender/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # reading a cell
        # print(worksheet["A1"].value)
        # print(worksheet["B1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row and printing it
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                # for every cell in a row, the first value is the name to be attached to the formatted string
                # carrying the message, while the second value which is in the second cell of the row is the
                # recipient at that particular instant
                row_data.append(str(cell.value))
                # print(cell.value)
            excel_data.append(row_data)
            print(row_data[0])
            print(row_data[1])

            name = (row_data[0])
            phone = int((row_data[1]))

            # sending the messages
            username = "testjoe"
            apikey = "065ce0c2c30f8bc4a3b038a2935b59b6b1baac0e8fd4a5d3103f7fb6149773ea"

            # Initialize the SDK
            africastalking.initialize(username, apikey)

            # Get SMS service
            sms = africastalking.SMS

            # Define some options to send the SMS
            recipients = ['+254719828205']
            message = 'I\'m {name}, and am all cool at night and I work all day at day'
            sender = '33334'

            # Send the SMS
            try:
                # Once this is done, that's it! We'll handle the rest
                response = sms.send(message, recipients, sender)
                print(response)
            except Exception as e:
                print(f"yoh cute ass nigger, we have a problem {e}")

        return render(request, 'smssender/index.html', {"excel_data": excel_data})
